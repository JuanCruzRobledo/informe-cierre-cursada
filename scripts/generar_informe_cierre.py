# -*- coding: utf-8 -*-
"""
Calcula el Estado Alumno (PROMOCIONA / REGULARIZA / RECURSA) y genera el Excel
final de cierre de cursada a partir de un JSON con los datos ya clasificados
del calificador de Moodle (ver references/moodle-export.md para cómo armarlo).

Uso:
    python generar_informe_cierre.py --input datos.json --output Cierre_Cursada.xlsx

Formato de entrada (--input), ver assets/ejemplo_input.json para un caso completo:

{
  "reglas_default": {
    "autoeval_min_pct": 90,
    "parcial_promocion_min_pct": 60,
    "parcial_regulariza_min_pct": 40,
    "tpi_min_pct": 60
  },
  "materias": [
    {
      "nombre_sheet": "Programacion 1",
      "umbral_tp_pct": 100,
      "reglas": {},                    // opcional, sobreescribe reglas_default para esta materia
      "tp_labels": ["TP 1", "TP 2", ...],   // opcional, default "TP 1".."TP N"
      "comisiones": [
        {
          "titulo": "PROGRAMACION 1 - COMISION 1 - TUTOR MATIAS SANTIAGO TORRES",
          "alumnos": [
            {
              "nombre": "GABRIEL ALFREDO BRIZUELA JIMENEZ",
              "email": "gabojimenez140@gmail.com",
              "tps": ["Aprobado", "Aprobado", "Desaprobado", "N/E", ...],
              "autoeval_pcts": [96.3, 100.0, null, ...],   // % por unidad, null = no rendida
              "parcial1_pcts": [60, null],                 // una entrada por instancia rendida
              "parcial2_pcts": [100],
              "tpi_pcts": [90]
            }
          ]
        }
      ]
    }
  ]
}

Los "Cuestionario de Actividad 1/2/3" (miniquizzes de práctica dentro de cada
unidad) NO entran en este JSON — se descartan en el paso de clasificación
(Fase 3 de SKILL.md), antes de llegar acá. Ver references/reglas-cierre.md.
"""
import argparse
import json
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_REGLAS = {
    "autoeval_min_pct": 90,
    "parcial_promocion_min_pct": 60,
    "parcial_regulariza_min_pct": 40,
    "tpi_min_pct": 60,
}

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TITULO_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
TITULO_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FONT = Font(bold=True)
RESUMEN_FONT = Font(bold=True)


def round_half_up(x):
    return math.floor(x + 0.5)


def max_or_none(valores):
    limpios = [v for v in (valores or []) if v is not None]
    return max(limpios) if limpios else None


def calcular_estado(alumno, reglas, umbral_tp_pct):
    """Devuelve el veredicto de cierre para un alumno según el árbol de decisión
    acordado: autoeval >=90% EN TODAS las unidades (no promedio), % de TPs
    aprobados >= umbral configurado, y la nota más alta entre TODAS las
    instancias rendidas de cada parcial y del TPI."""
    tps = alumno.get("tps", [])
    total_tps = len(tps)
    tps_aprobados = sum(1 for t in tps if t == "Aprobado")
    tp_ok = total_tps == 0 or (tps_aprobados / total_tps * 100) >= umbral_tp_pct

    autoeval_pcts = alumno.get("autoeval_pcts", [])
    autoeval_ok = len(autoeval_pcts) > 0 and all(
        v is not None and v >= reglas["autoeval_min_pct"] for v in autoeval_pcts
    )

    p1_max = max_or_none(alumno.get("parcial1_pcts"))
    p2_max = max_or_none(alumno.get("parcial2_pcts"))
    tpi_max = max_or_none(alumno.get("tpi_pcts"))

    promociona = (
        autoeval_ok
        and tp_ok
        and p1_max is not None and p1_max >= reglas["parcial_promocion_min_pct"]
        and p2_max is not None and p2_max >= reglas["parcial_promocion_min_pct"]
        and tpi_max is not None and tpi_max >= reglas["tpi_min_pct"]
    )
    regulariza = (not promociona) and (
        autoeval_ok
        and tp_ok
        and p1_max is not None and p1_max >= reglas["parcial_regulariza_min_pct"]
        and p2_max is not None and p2_max >= reglas["parcial_regulariza_min_pct"]
    )

    if promociona:
        nf_cruda = ((p1_max + p2_max) / 2) * 0.4 + tpi_max * 0.6
        nota_final = round_half_up(nf_cruda / 10)
        habilitado_final = "No aplica (promocionó)"
        estado = "PROMOCIONA"
    elif regulariza:
        nota_final = "N/E"
        habilitado_final = "Sí" if (tpi_max is not None and tpi_max >= reglas["tpi_min_pct"]) else "No (falta aprobar TPI)"
        estado = "REGULARIZA"
    else:
        nota_final = "N/E"
        habilitado_final = "No"
        estado = "RECURSA"

    return {
        "estado": estado,
        "nota_final": nota_final,
        "habilitado_final": habilitado_final,
        "p1_max": p1_max if p1_max is not None else "N/E",
        "p2_max": p2_max if p2_max is not None else "N/E",
        "tpi_max": tpi_max if tpi_max is not None else "N/E",
    }


def escribir_materia(wb, materia):
    nombre_sheet = materia["nombre_sheet"][:31]  # límite de Excel para nombres de hoja
    ws = wb.create_sheet(title=nombre_sheet)

    reglas = {**DEFAULT_REGLAS, **materia.get("reglas", {})}
    umbral_tp_pct = materia["umbral_tp_pct"]

    veredictos_por_comision = []
    for comision in materia["comisiones"]:
        veredictos = [calcular_estado(a, reglas, umbral_tp_pct) for a in comision["alumnos"]]
        veredictos_por_comision.append(veredictos)

    todos_veredictos = [v for grupo in veredictos_por_comision for v in grupo]
    promocionados = sum(1 for v in todos_veredictos if v["estado"] == "PROMOCIONA")
    regulares = sum(1 for v in todos_veredictos if v["estado"] == "REGULARIZA")
    recursantes = sum(1 for v in todos_veredictos if v["estado"] == "RECURSA")

    ws.append([f"TOTAL MATERIA {materia['nombre_sheet'].upper()}"])
    ws["A1"].font = RESUMEN_FONT
    ws.append(["PROMOCIONADOS", promocionados])
    ws.append(["REGULARES", regulares])
    ws.append(["RECURSANTES", recursantes])
    ws.append([])

    primer_alumno = next(
        (a for c in materia["comisiones"] for a in c["alumnos"]), None
    )
    n_tps = len(primer_alumno["tps"]) if primer_alumno else 0
    tp_labels = materia.get("tp_labels") or [f"TP {i + 1}" for i in range(n_tps)]

    headers = (
        ["Nombre y Apellido", "Email"]
        + tp_labels
        + ["Parcial 1", "Parcial 2", "TPI", "Nota Final", "Estado Alumno", "Habilitado para Final"]
    )

    for comision, veredictos in zip(materia["comisiones"], veredictos_por_comision):
        titulo_row = ws.max_row + 1
        ws.append([comision["titulo"]])
        ws.merge_cells(start_row=titulo_row, start_column=1, end_row=titulo_row, end_column=len(headers))
        celda_titulo = ws.cell(row=titulo_row, column=1)
        celda_titulo.fill = TITULO_FILL
        celda_titulo.font = TITULO_FONT
        celda_titulo.alignment = Alignment(horizontal="center")

        ws.append(headers)
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            celda = ws.cell(row=header_row, column=col)
            celda.font = HEADER_FONT
            celda.fill = HEADER_FILL

        for alumno, veredicto in zip(comision["alumnos"], veredictos):
            fila = (
                [alumno["nombre"], alumno["email"]]
                + list(alumno.get("tps", []))
                + [
                    veredicto["p1_max"],
                    veredicto["p2_max"],
                    veredicto["tpi_max"],
                    veredicto["nota_final"],
                    veredicto["estado"],
                    veredicto["habilitado_final"],
                ]
            )
            ws.append(fila)

        ws.append([])

    for col in range(1, len(headers) + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 18 if col > 2 else 28


def main():
    ap = argparse.ArgumentParser(description="Genera el Excel de cierre de cursada.")
    ap.add_argument("--input", required=True, help="JSON con los datos clasificados (ver docstring)")
    ap.add_argument("--output", required=True, help="Ruta del .xlsx final")
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        data = json.load(f)

    global DEFAULT_REGLAS
    DEFAULT_REGLAS = {**DEFAULT_REGLAS, **data.get("reglas_default", {})}

    wb = Workbook()
    wb.remove(wb.active)  # la hoja default en blanco

    for materia in data["materias"]:
        escribir_materia(wb, materia)

    wb.save(args.output)
    print(f"Informe generado: {args.output}")


if __name__ == "__main__":
    main()
